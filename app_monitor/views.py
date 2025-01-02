from django.shortcuts import render, HttpResponse,redirect
from .models import Temperature,Device,Alarm,Event
from django.views.decorators.csrf import csrf_exempt,requires_csrf_token
from django.utils import timezone
from django.core.paginator import Paginator
import django_excel as excel
# from django.http import HttpResponse
from openpyxl import Workbook
from django.db.models import Q
from django.db.models.functions import Lower #241029
from django.conf import settings #241105
from datetime import datetime
#seri-port-usb-config
import serial
import serial.tools.list_ports
import time
from django.contrib import messages #241115
import json #241120
from itertools import chain #241229 querysetleri birleştirmek için.
from operator import attrgetter # 241229 sorted() fonksiyonu için 


# from django.contrib

# Create your views here.

def ilk_def(request):
    return HttpResponse("Monitor edilecek sayfadasınız...")

def dashboard(request):
    print("dashboard girdi....")
    # device_query=Device.objects.all().order_by('device_name')
    device_query=Device.objects.all().order_by(Lower('device_name'))
    deviceAll = Temperature.objects.order_by('-id')
    paginator = Paginator(deviceAll, 10)  # Show 5 contacts per page.

    device_search_count = deviceAll.count()

    page_number = request.GET.get('page')

    devicePaginator = paginator.get_page(page_number)
    # device_query = paginator.get_page(page_number)
    context=dict(
        device_query=device_query,
        devicePaginator=devicePaginator,
    )
    return render(request,"app_monitor/dashboard.html",context)

# @csrf_exempt
# @requires_csrf_token
def tempList(request):
    tempsajax = Temperature.objects.order_by('-id')[:10]
    deviceAll = Temperature.objects.order_by('-id')
    kayit_sayisi_total_ajax = Temperature.objects.count()

    cihaz_isimleri_query=Temperature.objects.values("device_name").distinct()
    cihaz_isimleri=[]
    print('Temperature.objects.values("device_name")')
    print(cihaz_isimleri_query)
    for cihaz in cihaz_isimleri_query:
        cihaz_isimleri.append(list(cihaz.values())[0]) #cihaz isimlerini dizi olarak elde ettik.
    print(f"cihaz isimleri dizi: {cihaz_isimleri}")
    print("Templist girdi....")
    paginator = Paginator(deviceAll, 10)  # Show 10 contacts per page.

    device_search_count = deviceAll.count()

    page_number = request.GET.get('page')

    devicePaginator = paginator.get_page(page_number)
    tempsajax = paginator.get_page(page_number)
    context=dict(
        tempsajax=tempsajax,
        devicePaginator=devicePaginator,
        kayit_sayisi_total_ajax=kayit_sayisi_total_ajax,
    )
    return render(request, 'app_monitor/temperature_ajax.html', context)

# @csrf_exempt
def TemperatureAddRecord(request): ####RANDOM SICAKLIK, BSC1 tabloları doldurma, form işlemi olmadan, fetchdata_perf() içinde ajax yöntemiyle.

    kayit = request.POST.get("SicaklikKayit")
    newRecord = Temperature(temperature=kayit, date=timezone.now())
    newRecord.save()

    tempsajax = Temperature.objects.order_by('-id')[:10]
    kayit_sayisi_total_ajax = Temperature.objects.count()

    context=dict(
        tempsajax=tempsajax,
        kayit_sayisi_total_ajax=kayit_sayisi_total_ajax,
    )
    return render(request, 'app_monitor/temperature_ajax.html', context)


def addRecordArduino(request): # yeni sıcaklık kaydı ekleme,form get metoduyla
    if request.method == "GET":
        kayit = request.GET.get("kayit")
        humidity = request.GET.get("humidity")
        volcum = request.GET.get("volcum") #voltaj degeri
        device_name=request.GET.get("device_name")
        cikis1=request.GET.get("cikis1")
        print(f"Entry.DoesNotExist......??: {Device.objects.filter(device_id=3)}")
        # if not Device.objects.filter(device_name__icontains=device_name).exists():
        if not Device.objects.filter(device_name__iexact=device_name).exists():
            # temp = Temperature.objects.filter(device_name__iexact=cihazadi) #iexact kullanıldı exact değil
            print(f" {device_name}:  device_id database de olmayan blok girdi... ")
            device_id_request=request.GET.get("device_id")
            device_ip_request=request.GET.get("device_ip")
            device_port_request=request.GET.get("device_port")            
            new_device=Device(device_id=device_id_request,device_name=device_name,device_port=device_port_request,device_ip=device_ip_request)
            new_device.save()
            print(f"new_device: {new_device}")
            print(f"new device name: {new_device.device_name}")

        device_id_request=request.GET.get("device_id")
        if not Device.objects.filter(device_id=device_id_request).exists():
            print(f" {device_id_request}:  device_id database de olmayan blok girdi... ")
            # device_id_request=request.GET.get("device_id")
            device_ip_request=request.GET.get("device_ip")
            device_port_request=request.GET.get("device_port")
            new_device=Device(device_id=device_id_request,device_name=device_name,device_port=device_port_request,device_ip=device_ip_request)
            new_device.save()
            print(f"new_device: {new_device}")
            print(f"new device name: {new_device.device_name}")
        device_port_request=request.GET.get("device_port") 
        device_check=Device.objects.get(device_id=device_id_request)
        if device_check.device_port !=int(device_port_request):
            device_check.device_port=int(device_port_request)
            device_check.save()
        
        # device_id=Device.objects.get(device_name__icontains=device_name) #241013, device_id Device örneği olmalı, Foreignkey
        device_id_request=request.GET.get("device_id")
        device_id=Device.objects.filter(device_name__icontains=device_name).get(device_id=device_id_request) #241013, device_id Device örneği olmalı, Foreignkey
        
        print(f"kayit: {kayit} - device_name: {device_name}")

        newRecord = Temperature(temperature=kayit,humidity=humidity ,volcum=volcum,device_name=device_name,device_id=device_id, date=timezone.now(),cikis1=cikis1) #241013
        newRecord.save()
        # newRecord = Temperature(temperature=kayit,humidity=humidity ,volcum=volcum,device_name=device_name, date=timezone.now())
    else:
        kayit = request.POST.get("kayit")
        if 50<=int(kayit)<=55: # Sıcaklık 50-55 arasında olduğunda mail gönderimi de yapılacaktır.
            newRecord = Temperature(temperature=kayit, date=timezone.now(),mailSend=True)
            email = EmailMessage('Sıcaklık Değeri!', 'Ölçülen Sıcaklık Değeri: ' + f"{kayit}" +' derecedir. \n (50-55 arası sıcaklıklar mail ile bilgilendirilir)', to=['uyarbaris@gmail.com'])
            email.send()
            messages.info(request,f"uyarbaris@gmail.com adresine mail gönderimi yapılmıştır. <br> Sıcaklık:{kayit} ")
        else:
            newRecord = Temperature(temperature=kayit, date=timezone.now())
    newRecord.save()
    print(f"son eklenen sıcaklık: {newRecord.temperature}")
    
    tempsajax = Temperature.objects.order_by('-id')[:10]
    kayit_sayisi_total_ajax = Temperature.objects.count()

    context=dict(
        tempsajax=tempsajax,
        kayit_sayisi_total_ajax=kayit_sayisi_total_ajax,

    )
    # return render(request, 'app_monitor/temperature_ajax.html', context)
    return redirect('/app_monitor')

# GIT commit 240921-2
# @csrf_exempt
# def deviceView(request,str_device_name):
def deviceView(request,str_device_name,port_no):
# def deviceView(request,str_device_name): # parametrelerin sırası ÖNEMLİ
    deviceAll=Temperature.objects.filter(device_name__iexact=str_device_name).order_by('-id')
    device=Temperature.objects.filter(device_name__iexact=str_device_name).order_by('-id')[:10]
    # device500=Temperature.objects.filter(device_name=str_device_name).order_by('-id')[:500]
    device500=Temperature.objects.filter(device_name__iexact=str_device_name).order_by('-id')[:500]
    # device_port=Device.objects.get(device_name=str_device_name).device_port
    device_port=port_no
    device_id_dizi=[]
    print(f"deviceView girdi, device={str_device_name} ")
    print(f"device çıktısı: , {device} ")
    print(f"str_device_name çıktısı: , {str_device_name} ")
    print(f"port no url parametre: , {port_no} ")
    paginator = Paginator(deviceAll, 5)  # Show 5 contacts per page.

    device_search_count = deviceAll.count()

    page_number = request.GET.get('page')

    devicePaginator = paginator.get_page(page_number)
    device_id_dizi=[device.device_id for device in Device.objects.filter(device_name=str_device_name)] #241106 List Comprehension ile dizi oluşturma
    print(f"{str_device_name} device ID ler: {device_id_dizi}")
    context=dict(
        device=device,
        device_name=str_device_name.lower(),
        device500=device500,
        devicePaginator=devicePaginator,
        device_search_count=device_search_count,
        device_port=device_port,
        device_id_dizi=device_id_dizi,
    )
    return render(request,"app_monitor/device.html",context)


def deviceViewDetail(request,str_device_name,port_no,device_id): # parametrelerin sırası ÖNEMLİ
    deviceAll=Temperature.objects.filter(device_name=str_device_name).filter(device_id=device_id).order_by('-id')
    device=Temperature.objects.filter(device_name=str_device_name).filter(device_id=device_id).order_by('-id')[:10]
    # device500=Temperature.objects.filter(device_name=str_device_name).order_by('-id')[:500]
    device500=Temperature.objects.filter(device_name=str_device_name).filter(device_id=device_id).order_by('-id')[:50]
    # device_port=Device.objects.get(device_name=str_device_name).device_port
    device_port=port_no
    print(f"deviceView girdi, device={str_device_name} ")
    print(f"device çıktısı: , {device} ")
    print(f"str_device_name çıktısı: , {str_device_name} ")
    print(f"port no url parametre: , {port_no} ")
    paginator = Paginator(deviceAll, 5)  # Show 5 contacts per page.

    device_search_count = deviceAll.count()

    page_number = request.GET.get('page')

    devicePaginator = paginator.get_page(page_number)

    context=dict(
        device=device,
        device_name=str_device_name.lower(),
        device500=device500,
        devicePaginator=devicePaginator,
        device_search_count=device_search_count,
        device_port=device_port,
    )
    return render(request,"app_monitor/device.html",context)

def device_id(request,device_id): #241111
# def deviceView(request,str_device_name): # parametrelerin sırası ÖNEMLİ
    device_id_query=Temperature.objects.filter(device_id=device_id).order_by('-id')

    paginator = Paginator(device_id_query, 5)  # Show 5 contacts per page.

    device_search_count = device_id_query.count()

    page_number = request.GET.get('page')

    devicePaginator = paginator.get_page(page_number)
    # device_id_dizi=[device.device_id for device in Device.objects.filter(device_name=str_device_name)] #241106 List Comprehension ile dizi oluşturma
    # print(f"{str_device_name} device ID ler: {device_id_dizi}")
    context=dict(
        devicePaginator=devicePaginator,
        device_search_count=device_search_count,
        device_id=str(device_id),
        # device_port=device_port,

    )
    return render(request,"app_monitor/device_id.html",context)

#Excel export
def exportExcel(request):
    temps10=Temperature.objects.order_by('-id')[:10]
    
    return excel.make_response_from_a_table(Temperature, "xls", file_name="sheet")

#Excel 
def export_to_excel_serial_query(request):
    id1=request.GET.get("id1")
    id2=request.GET.get("id2")
    sicaklik1=request.GET.get("sicaklik1")
    sicaklik2=request.GET.get("sicaklik2")
    nem1=request.GET.get("nem1")
    nem2=request.GET.get("nem2")
    voltaj1=request.GET.get("voltaj1")
    voltaj2=request.GET.get("voltaj2")
    tarih1=request.GET.get("tarih1")
    tarih2=request.GET.get("tarih2")   
    cihazadi=request.GET.get("cihazadi")

    print(f"get id1 to excel:{id1}")
    print(f"get id2 to excel:{id2}")

    #ID aralığı
    if (id1=="" or None) and (id2=="" or None):
        filter_result_id=Temperature.objects.filter(
            id__gte=1,id__lte=Temperature.objects.last().id
            ).filter(device_name=cihazadi.capitalize()).order_by('-id')
        print(f"(id1=='' or None) and (id2==' or None) sayısı:{filter_result_id.count()}")
    elif id1=="" or None: 
        filter_result_id=Temperature.objects.filter(
            id__gte=1,id__lte=id2
            ).filter(device_name=cihazadi.capitalize()).order_by('-id')
        print(f"id1=='' or None")
    elif id2=="" or None:
        filter_result_id=Temperature.objects.filter(
            id__gte=id1
            ).filter(device_name=cihazadi.capitalize()).order_by('-id')
        print(f"id2=='' or None")
    
    else:
        filter_result_id=Temperature.objects.filter(
                id__gte=id1,id__lte=id2
                ).filter(device_name=cihazadi.capitalize()).order_by('-id')
        print(f"id__gte=id1,id__lte=id2")
        print(f"id sayısı: {filter_result_id.count()}")
    
    #SICAKLIK aralığı
    if (sicaklik1=="" or None) and (sicaklik2=="" or None): 
            filter_result_temp=filter_result_id.filter(
            # temperature__gte=0,temperature__lte=sicaklik2
            Q(temperature__gte=0,temperature__lte=100)|
            Q(temperature__exact=None)
            ).filter(device_name__iexact=cihazadi).order_by('-id')  
            print(f"sicaklik kontrolu: sicaklik1=='' or None and sicaklik2=='' or None")  
            print(f"temp sayısı none-none: {filter_result_temp.count()}")   
    elif sicaklik1=="" or None: #SICAKLIK aralığı
            filter_result_temp=filter_result_id.filter(
            # temperature__gte=0,temperature__lte=sicaklik2
            temperature__gte=0,temperature__lte=sicaklik2
            ).filter(device_name__iexact=cihazadi).order_by('-id')       
    elif sicaklik2=="" or None:
        filter_result_temp=filter_result_id.filter(
                temperature__gte=sicaklik1
                ).filter(device_name__iexact=cihazadi).order_by('-id')
    else:
            filter_result_temp=filter_result_id.filter(
            temperature__gte=sicaklik1,temperature__lte=sicaklik2
            ).filter(device_name__iexact=cihazadi).order_by('-id')
    #NEM aralığı
    if (nem1=="" or None) and (nem2=="" or None): 
        filter_result_nem=filter_result_temp.filter(
            Q(humidity__gte=0,humidity__lte=100)|
            Q(humidity__exact=None)
            ).filter(device_name__iexact=cihazadi).order_by('-id')
        print(f"nem kontrolu:nem1=="" or None) and (nem2=="" or None ")
        print(f"nem sayısı none-none: {filter_result_nem.count()}") 
    elif nem1=="" or None: 
        filter_result_nem=filter_result_temp.filter(
            humidity__gte=0,humidity__lte=nem2
            ).filter(device_name__iexact=cihazadi).order_by('-id')
    elif nem2=="" or None:
        filter_result_nem=filter_result_temp.filter(
            humidity__gte=nem1
            ).filter(device_name__iexact=cihazadi).order_by('-id')
    else:
        filter_result_nem=filter_result_temp.filter(
                humidity__gte=nem1,humidity__lte=nem2
                ).filter(device_name=cihazadi.capitalize()).order_by('-id')  
        print(f"nem kontrolu:humidity__gte=nem1,humidity__lte=nem2") 
        print(f"nem sayısı nem1-nem2: {filter_result_nem.count()}")  
    #VOLTAJ aralığı
    if (voltaj1=="" or None) and (voltaj2=="" or None): 
        filter_result_voltaj=filter_result_nem.filter(
            volcum__gte=0,volcum__lte=20
            ).filter(device_name__iexact=cihazadi).order_by('-id')
        print(f"voltaj kontrolu:voltaj1=="" or None) and (voltaj2=="" or None ")
        print(f"voltaj sayısı none-none: {filter_result_voltaj.count()}") 
    elif voltaj1=="" or None: 
        filter_result_voltaj=filter_result_nem.filter(
            volcum__gte=0,volcum__lte=voltaj2
            ).filter(device_name__iexact=cihazadi).order_by('-id')
    elif voltaj2=="" or None:
        filter_result_voltaj=filter_result_nem.filter(
            volcum__gte=voltaj1
            ).filter(device_name__iexact=cihazadi).order_by('-id')
    else:
        filter_result_voltaj=filter_result_nem.filter(
                volcum__gte=voltaj1,volcum__lte=voltaj2
                ).filter(device_name__iexact=cihazadi).order_by('-id')    
        print(f"voltaj sayısı: {filter_result_voltaj.count()}")
        print(f"nem sayısı: {filter_result_nem.count()}")

    #TARIH aralığı
    # print(f"tarih1 tipi: {type(tarih1)}")
    # print(f"tarih1: {tarih1}")
    # #2024-11-12 22:40:06.395707   '%Y-'
    # if (tarih1!="" or "None") and (tarih2!="" or "None"): 
    #     tarih1_datetime=datetime.strptime(tarih1,'%Y-%m-%dT%H:%M') #string-datetime
    #     tarih1=datetime.strftime(tarih1_datetime,'%Y-%m-%d %H:%M') #datetime format değiştirme
    #     tarih2_datetime=datetime.strptime(tarih2,'%Y-%m-%dT%H:%M') #string-datetime
    #     tarih2=datetime.strftime(tarih2_datetime,'%Y-%m-%d %H:%M') #datetime format değiştirme
    #     # tarih1_datetime=datetime.date(tarih1)
    #     # print(f"tarih1_datetime tipi: {type(tarih1_datetime)}")
    # if (tarih1=="" or "None") and (tarih2=="" or "None"): 
    #     filter_result_tarih=filter_result_voltaj.filter(
    #         date__gte=datetime(2023,12,30),date__lte=datetime.now()
    #         ).filter(device_name__iexact=cihazadi).order_by('-id')
    #     print(f"tarih kontrolu:tarih1=="" or None) and (tarih2=="" or None ")
    #     print(f"tarih sayısı none-none: {filter_result_tarih.count()}") 
    # elif tarih1=="" or None: 
    #     filter_result_tarih=filter_result_voltaj.filter(
    #         date__lte=tarih2
    #         ).filter(device_name__iexact=cihazadi).order_by('-id')
    # elif tarih2=="" or None:
    #     filter_result_tarih=filter_result_voltaj.filter(
    #         date__gte=tarih1
    #         ).filter(device_name__iexact=cihazadi).order_by('-id')
    # else:
    #     filter_result_tarih=filter_result_voltaj.filter(
    #             date__gte=tarih1,date__lte=tarih2
    #             ).filter(device_name__iexact=cihazadi).order_by('-id')    
    #     print(f"tarih sayısı: {filter_result_tarih.count()}")

    print(f"cihazadi export_to_excel= {cihazadi}")
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{cihazadi}.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = cihazadi

    # Add headers
    headers = ["id","DEVICE NAME","DEVICE ID","DEVICE PORT","temperature", "humidity", "volcum","TARIH","ACIKLAMA"]
    ws.append(headers)

    # Add data from the model
    temp=filter_result_voltaj
    print(f"filter_result_voltaj:{filter_result_voltaj}")
    for temps in temp:
        ws.append([temps.id,temps.device_name,temps.device_id.device_id,temps.device_id.device_port,temps.temperature, temps.humidity, temps.volcum,temps.date,temps.additionalText])

    # Save the workbook to the HttpResponse
    wb.save(response)
    return response

#Excel export 2
def export_to_excel(request):
    cihazadi=request.GET.get("cihazadi")
    id1=request.GET.get("id1")
    id2=request.GET.get("id2")
    nem1=request.GET.get("nem1")
    nem2=request.GET.get("nem2")
    print(f"cihazadi export_to_excel= {cihazadi}")
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{cihazadi}.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = cihazadi

    # Add headers
    headers = ["id","DEVICE NAME","DEVICE ID","DEVICE PORT","temperature", "humidity", "volcum","TARIH","ACIKLAMA"]
    ws.append(headers)

    # Add data from the model
    if (nem1 or nem2) != "" or None:
        # temp = Temperature.objects.filter(device_name__icontains=cihazadi,humidity__gte=nem1,humidity__lte=nem2)
        temp = Temperature.objects.filter(device_name=cihazadi,humidity__gte=nem1,humidity__lte=nem2)
    elif (id1 or id2) != "" or None:
        temp = Temperature.objects.filter(device_name=cihazadi,id__gte=id1,id__lte=id2)
        print(f"excel id filtre sonuç nesneleri: {temp}")
    else:
        # temp = Temperature.objects.filter(device_name__icontains=cihazadi)
        temp = Temperature.objects.filter(device_name__iexact=cihazadi) #iexact kullanıldı exact değil
        # temp = Temperature.objects.filter(device_name=cihazadi)
    for temps in temp:
        ws.append([temps.id,temps.device_name,temps.device_id.device_id,temps.device_id.device_port,temps.temperature, temps.humidity, temps.volcum,temps.date,temps.additionalText])

    # Save the workbook to the HttpResponse
    wb.save(response)
    return response

#Excel export ALL
def export_to_excel_all(request):
    # cihazadi=request.GET.get("cihazadi")
    # id1=request.GET.get("id1")
    # id2=request.GET.get("id2")
    # nem1=request.GET.get("nem1")
    # nem2=request.GET.get("nem2")
    # print(f"cihazadi export_to_excel= {cihazadi}")
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="ALL.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "ALL"

    # Add headers
    headers = ["id","DEVICE NAME","DEVICE ID","DEVICE PORT","temperature", "humidity", "volcum","TARIH","ACIKLAMA"]
    ws.append(headers)


    temp = Temperature.objects.all() 
        # temp = Temperature.objects.filter(device_name=cihazadi)
    for temps in temp:
        try:
            ws.append([temps.id,temps.device_name,temps.device_id.device_id,temps.device_id.device_port,temps.temperature, temps.humidity, temps.volcum,temps.date,temps.additionalText])
            # ws.append([temps.id,temps.device_name,temps.temperature, temps.humidity, temps.volcum])
        except AttributeError:
            print("AttributeError oluştu...")

    # Save the workbook to the HttpResponse
    wb.save(response)
    return response

#Excel export ID
def export_to_excel_id(request):
    cihazid=request.GET.get("cihazid")
    # id1=request.GET.get("id1")
    # id2=request.GET.get("id2")
    # nem1=request.GET.get("nem1")
    # nem2=request.GET.get("nem2")
    # print(f"cihazadi export_to_excel= {cihazadi}")
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{cihazid}.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = f"{cihazid}"

    # Add headers
    headers = ["id","DEVICE NAME","DEVICE ID","DEVICE PORT","temperature", "humidity", "volcum","tarih"]
    ws.append(headers)


    temp = Temperature.objects.filter(device_id=cihazid) 
        # temp = Temperature.objects.filter(device_name=cihazadi)
    for temps in temp:
        try:
            ws.append([temps.id,temps.device_name,temps.device_id.device_id,temps.device_id.device_port,temps.temperature, temps.humidity, temps.volcum,str(temps.date)])
            # ws.append([temps.id,temps.device_name,temps.temperature, temps.humidity, temps.volcum])
        except AttributeError:
            print("AttributeError oluştu...")

    # Save the workbook to the HttpResponse
    wb.save(response)
    return response


def export_to_excel_nem(request):
    nem1=request.GET.get("nem1")
    nem2=request.GET.get("nem2")
    cihazadi=request.GET.get("cihazadi")
    print(f"cihazadi= {cihazadi}")
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{cihazadi}.xlsx"'
    wb = Workbook()
    ws = wb.active
    ws.title = cihazadi
    headers = ["device_name","id","temperature", "humidity", "volcum"]
    ws.append(headers)
    temp = Temperature.objects.filter(device_name__icontains=cihazadi,humidity__gte=nem1,humidity__lte=nem2)
    for temps in temp:
        ws.append([temps.device_name,temps.id,temps.temperature, temps.humidity, temps.volcum])
    wb.save(response)
    return response

def django_device(request):
    device_ip = request.GET.get("device_ip")
    device_port=request.GET.get("name-port")
    device_name=request.GET.get("device_name")
    print(f"django_device girdi: {device_ip}")
    devices_all=Device.objects.all()
    cihazlar_erisim_ip=settings.CIHAZLAR_ERISIM_IP

    context=dict(
        device_ip=device_ip,
        device_port=device_port,
        devices_all=devices_all,
        cihazlar_erisim_ip=cihazlar_erisim_ip,
        device_name=device_name,
    )

    return render(request,"app_monitor/django_arduino.html",context)

def django_device_backtest(request):
    device_ip = request.GET.get("device_ip")
    print(f"DJANGO BACKTEST girdi: {device_ip}")
    return HttpResponse(device_ip)

def devices_all(request):
    devices_all=Device.objects.all()
    datetime_now=datetime.now()
    devices_online=[]
    cihaz_isimleri_query=Temperature.objects.values("device_name").distinct()
    cihaz_isimleri=[]
    for cihaz in cihaz_isimleri_query:
        cihaz_isimleri.append(list(cihaz.values())[0]) #cihaz isimlerini dizi olarak elde ettik. 250101

    print(f"devices_all:{devices_all}")
    # print(f"datetime.now(){datetime_now}")
    # print(f"timestamp now {datetime.timestamp(datetime_now)}")
    for device in devices_all:
        if datetime.timestamp(datetime_now) - datetime.timestamp(device.temperature_set.last().date) < 360:
            devices_online.append(device.temperature_set.last().device_id.device_id)

            print(f"datetime_now- device.temperature_set.last.date() {device.temperature_set.last().device_id}: {datetime.timestamp(datetime_now) - datetime.timestamp(device.temperature_set.last().date) }")
    #     if timezone.now - device.temperature_set.last.date 
    print(f"online cihazlar: {devices_online}")
    context=dict(
        devices_all=devices_all,
        devices_online=devices_online,
        cihaz_isimleri=cihaz_isimleri,

    )
    return render(request,"app_monitor/devices_all.html",context)

######################USB-SERI-PORT-CONFIG######################
def arduino_serial_local(request,config_parameter):
    comport=request.GET.get("comport")
    id=request.GET.get("cihaz-id")
    name=request.GET.get("cihaz-adi")
    serverip=request.GET.get("server-ip")
    cihazip=request.GET.get("cihaz-ip")
    ssid=request.GET.get("cihaz-ssid")
    password=request.GET.get('cihaz-password')
    cihazport=request.GET.get("cihaz-port")
    gatewayip=request.GET.get('default-gateway')
    print(f"comport:{comport}")
    config_parameter=config_parameter
    if config_parameter == "all":
        value  = write_read(config_parameter) #241117
        print(f"arduino value: {value}, type: {type(value)}")
        port_listesi=[]
        all_serial_ports=serial.tools.list_ports.comports()
        for port in list(serial.tools.list_ports.comports()):
            print(port[0])
            port_listesi.append(port[0])
        print(f"port listesi: {port_listesi}")
        myports = [tuple(p) for p in list(serial.tools.list_ports.comports())]
        print(myports)
        print(f"all serial ports: {all_serial_ports}")
        try:
            value_dict=json.loads(value)

            context=dict(
            value=value,
            device_id=value_dict['CihazId'],
            device_name=value_dict['CihazAdi'],
            device_port=value_dict['CihazPort'],
            device_ip=value_dict['CihazIp'],
            server_ip=value_dict['ServerIp'],
            device_ssid=value_dict['CihazSSID'],
            device_password=value_dict['CihazPassword'],
            ag_gecidi=value_dict['AgGecidi'],
            myports=myports,
            port_listesi=port_listesi,
        )
        except:
            context=dict(
            value="USB Portu kontrol ediniz",
            device_id="USB Portu kontrol ediniz",
            device_name="USB Portu kontrol ediniz",
            device_port="USB Portu kontrol ediniz",
            device_ip="USB Portu kontrol ediniz",
            server_ip="USB Portu kontrol ediniz",
            device_ssid="USB Portu kontrol ediniz",
            device_password="USB Portu kontrol ediniz",
            ag_gecidi="USB Portu kontrol ediniz",
            myports=myports,
            port_listesi=port_listesi,
        )

    elif config_parameter == "id":
        value  = write_read_id(id) #241117
        print(f"arduino value: {value}, type: {type(value)}")
        value_dict=json.loads(value)
        context=dict(
        value=value,
        device_id=value_dict['CihazId'],
        )
    elif config_parameter== "name":
        value = write_read_name(name)
        print(f"arduino value: {value}, type: {type(value)}")
        value_dict=json.loads(value)
        context=dict(
        value=value,
        device_name=value_dict['CihazAdi'],
        )
    elif config_parameter== "serverip":
        value = write_read_serverip(serverip)
        print(f"arduino value: {value}, type: {type(value)}")
        value_dict=json.loads(value)
        context=dict(
        value=value,
        server_ip=value_dict['ServerIp'],
        )       
    elif config_parameter== "cihazip":
        value = write_read_cihazip(cihazip)
        print(f"arduino value: {value}, type: {type(value)}")
        value_dict=json.loads(value)
        context=dict(
        value=value,
        device_ip=value_dict['CihazIp'],
        )    
    elif config_parameter== "gatewayip":
        value = write_read_gatewayip(gatewayip)
        print(f"arduino value: {value}, type: {type(value)}")
        value_dict=json.loads(value)
        context=dict(
        value=value,
        ag_gecidi=value_dict['AgGecidi'],
        )
    elif config_parameter== "ssid":
        value = write_read_ssid(ssid)
        print(f"arduino value: {value}, type: {type(value)}")
        value_dict=json.loads(value)
        context=dict(
        value=value,
        device_ssid=value_dict['CihazSSID'],
        )  
    elif config_parameter== "password":
        value = write_read_password(password)
        print(f"arduino value: {value}, type: {type(value)}")
        value_dict=json.loads(value)
        context=dict(
        value=value,
        device_password=value_dict['CihazPassword'],
        )
    elif config_parameter== "cihazport":
        value = write_read_cihazport(cihazport)
        print(f"arduino value: {value}, type: {type(value)}")
        value_dict=json.loads(value)
        context=dict(
        value=value,
        device_port=value_dict['CihazPort'],
        )        
    elif config_parameter== "reset":
        value = write_read_reset()
    # value  = write_read(comport,config_parameter) #241117
    # value  = write_read(comport,config_parameter,id) #241117
    # value  = write_read(config_parameter,id) #241123
    print(f"arduino value: {value}, type: {type(value)}")
    # print(f"arduino value CihazId: {value[0:10]}")
    ## try:
    #     value_dict=json.loads(value)
    #     # value_dict=json.loads(value).strip("'<>() ").replace('\'', '\"')
    #     # value_dict=json.dumps(value)
    #     print(f"value_dict:{value_dict},type:{type(value_dict)}")
    #     port_listesi=[]
    #     all_serial_ports=serial.tools.list_ports.comports()
    #     for port in list(serial.tools.list_ports.comports()):
    #         print(port[0])
    #         port_listesi.append(port[0])
    #     print(f"port listesi: {port_listesi}")
    #     myports = [tuple(p) for p in list(serial.tools.list_ports.comports())]
    #     # print(myports)
    #     # print(f"all serial ports: {all_serial_ports}")
    #     # try:
    #     context=dict(
    #         value=value,
    #         # comport=comport,
    #         device_id=value_dict['CihazId'],
    #         device_name=value_dict['CihazAdi'],
    #         device_port=value_dict['CihazPort'],
    #         device_ip=value_dict['CihazIp'],
    #         server_ip=value_dict['ServerIp'],
    #         device_ssid=value_dict['CihazSSID'],
    #         device_password=value_dict['CihazPassword'],
    #         ag_gecidi=value_dict['AgGecidi'],
    #         # reset_ok=value_dict['ResetOk'],

    #         # device_port=value_dict["device_port"],
    #         # device_port=value_dict[0:5],
    #         myports=myports,
    #         port_listesi=port_listesi,
    #         #port=port,
    #     )
    # # except KeyError:
    #     # print("Arduinodan eksik key geldi...")
    # except :
    #     print("Arduinodan eksik key geldi...")
    #     context=dict(
    #         device_id="Veri okunamadı",
    #         device_name="Veri okunamadı",
    #     )
    if value == None:
        messages.info(request,f"COM portunu hatalı girdiniz...<br>COM port değeri: {value}")

    print(f"value: {value}")
    return render(request,"app_monitor/arduino_serial.html",context)

# def write_read(*parametreler):
def write_read(config_parameter):
    an=datetime.now()
    datetime_send=datetime.strftime(an,'%c')
    print(f"write_read_all girdi...")
# try:
    # comport="COM5"
    comport=list(serial.tools.list_ports.comports())[0][0]
    print(f"comport: {comport}")
    # arduino = serial.Serial(port=comport,  baudrate=115200, timeout=1)
    arduino = serial.Serial(port=comport,  baudrate=9600, timeout=.5,stopbits=serial.STOPBITS_ONE)
    print(f"arduino: {arduino}")
    an=datetime.now()
    datetime_send=datetime.strftime(an,'%c')
    # print(f"datetimesen tipi: {type(datetime_send)}")
    # if config_parameter == "all":
    arduino.write(bytes("all",  'utf-8')) #butun config göster
    # arduino.write(bytes("asd",  'utf-8')) #butun config göster
    print("arduino.write all parametresiyle çalıştı,alt satırı burası...")
    time.sleep(0.05)
    time.sleep(3)
    data = arduino.readline().decode('utf_8')
    print(f"arduino.readline():{data}")
    arduino.close()  ##### CLOSE() yapılmazsa bazen hata alınıyor.......Expecting value: line 1 column 1 (char 0)
    return  data
# except :
    print("FileNotFoundError exception oluştu")
    # arduino.close() #241117
    return "Exception Hata..."
    # raise Exception("COM portu yanlış giriyorsunuz...")

def write_read_id(id):
    try:
        id=id
        comport=list(serial.tools.list_ports.comports())[0][0]
        arduino = serial.Serial(port=comport,  baudrate=9600, timeout=.1)
        print(f"arduino: {arduino}")
        arduino_write=dict(
            id=id
        )
        arduino_write_dumps=json.dumps(arduino_write)
        print(f"arduino_write_dumps:{arduino_write_dumps}, type:{type(arduino_write_dumps)}")
        # arduino_write_str=str(arduino_write)
        # print(f"arduino_write:{arduino_write_str}")
        # arduino.write(bytes("1",  'utf-8')) #butun config göster
        arduino.write(bytes(arduino_write_dumps,  'utf-8')) #butun config göster
        # arduino.write(bytes("id",  'utf-8')) #butun config göster
        time.sleep(3)    ######## BU SATIR OLMAZSA KESINLIKLE HATA VERIYOR!!!!!!!!!!!!!!!!!!!!!!!
        data = arduino.readline().decode('utf_8')
        print(f"arduino.readline():{data}")
        arduino.close()
        return  data
        # return arduino_write_str
        # return arduino_write_dumps
        # return  "sabit return data"
        # return  '{"CihazId":1,"CihazAdi":"M(t","CihazPort":90,"CihazIp":"(IP unset)","ServerIp":"192.168.43.130","AgGecidi":"192.168.1.1"}'
    except :
        print("FileNotFoundError exception oluştu")
        return "Exception Hata..."

def write_read_name(name):
    try:
        name=name
        comport=list(serial.tools.list_ports.comports())[0][0]
        arduino = serial.Serial(port=comport,  baudrate=9600, timeout=.1)
        print(f"arduino: {arduino}")
        arduino_write=dict(
            name=name
        )
        arduino_write_dumps=json.dumps(arduino_write)
        print(f"arduino_write_dumps:{arduino_write_dumps}, type:{type(arduino_write_dumps)}")
        # arduino_write_str=str(arduino_write)
        # print(f"arduino_write:{arduino_write_str}")
        # arduino.write(bytes("1",  'utf-8')) #butun config göster
        arduino.write(bytes(arduino_write_dumps,  'utf-8')) #butun config göster
        # arduino.write(bytes("id",  'utf-8')) #butun config göster
        time.sleep(3)    ######## BU SATIR OLMAZSA KESINLIKLE HATA VERIYOR!!!!!!!!!!!!!!!!!!!!!!!
        data = arduino.readline().decode('utf_8')
        print(f"arduino.readline():{data}")
        return  data
        # return arduino_write_str
        # return arduino_write_dumps
        # return  "sabit return data"
        # return  '{"CihazId":1,"CihazAdi":"M(t","CihazPort":90,"CihazIp":"(IP unset)","ServerIp":"192.168.43.130","AgGecidi":"192.168.1.1"}'
    except :
        print("FileNotFoundError exception oluştu")
        return "Exception Hata..."

def write_read_serverip(serverip):
    try:
        serverip=serverip
        serverip_array=serverip.split('.')
        print(f"serverip_array: {serverip_array}")
        comport=list(serial.tools.list_ports.comports())[0][0]
        arduino = serial.Serial(port=comport,  baudrate=9600, timeout=.1)
        print(f"arduino: {arduino}")
        arduino_write=dict(
            serverip=serverip,
            serverip_3=serverip_array[3],
            serverip_2=serverip_array[2],
            serverip_1=serverip_array[1],
            serverip_0=serverip_array[0],
        )
        arduino_write_dumps=json.dumps(arduino_write)
        print(f"arduino_write_dumps:{arduino_write_dumps}, type:{type(arduino_write_dumps)}")
        # arduino_write_str=str(arduino_write)
        # print(f"arduino_write:{arduino_write_str}")
        # arduino.write(bytes("1",  'utf-8')) #butun config göster
        arduino.write(bytes(arduino_write_dumps,  'utf-8')) #butun config göster
        # arduino.write(bytes("id",  'utf-8')) #butun config göster
        time.sleep(3)    ######## BU SATIR OLMAZSA KESINLIKLE HATA VERIYOR!!!!!!!!!!!!!!!!!!!!!!!
        data = arduino.readline().decode('utf_8')
        print(f"arduino.readline():{data}")
        return  data
        # return arduino_write_str
        # return arduino_write_dumps
        # return  "sabit return data"
        # return  '{"CihazId":1,"CihazAdi":"M(t","CihazPort":90,"CihazIp":"(IP unset)","ServerIp":"192.168.43.130","AgGecidi":"192.168.1.1"}'
    except :
        print("FileNotFoundError exception oluştu")
        return "Exception Hata..."

def write_read_cihazip(cihazip):
    try:
        cihazip=cihazip
        cihazip_array=cihazip.split('.')
        print(f"serverip_array: {cihazip_array}")
        comport=list(serial.tools.list_ports.comports())[0][0]
        arduino = serial.Serial(port=comport,  baudrate=9600, timeout=.1)
        print(f"arduino: {arduino}")
        arduino_write=dict(
            cihazip=cihazip,
            cihazip_3=cihazip_array[3],
            cihazip_2=cihazip_array[2],
            cihazip_1=cihazip_array[1],
            cihazip_0=cihazip_array[0],
        )
        arduino_write_dumps=json.dumps(arduino_write)
        print(f"arduino_write_dumps:{arduino_write_dumps}, type:{type(arduino_write_dumps)}")
        # arduino_write_str=str(arduino_write)
        # print(f"arduino_write:{arduino_write_str}")
        # arduino.write(bytes("1",  'utf-8')) #butun config göster
        arduino.write(bytes(arduino_write_dumps,  'utf-8')) #butun config göster
        # arduino.write(bytes("id",  'utf-8')) #butun config göster
        time.sleep(3)    ######## BU SATIR OLMAZSA KESINLIKLE HATA VERIYOR!!!!!!!!!!!!!!!!!!!!!!!
        data = arduino.readline().decode('utf_8')
        print(f"arduino.readline():{data}")
        return  data
        # return arduino_write_str
        # return arduino_write_dumps
        # return  "sabit return data"
        # return  '{"CihazId":1,"CihazAdi":"M(t","CihazPort":90,"CihazIp":"(IP unset)","ServerIp":"192.168.43.130","AgGecidi":"192.168.1.1"}'
    except :
        print("FileNotFoundError exception oluştu")
        return "Exception Hata..."

def write_read_gatewayip(gatewayip):
    try:
        gatewayip=gatewayip
        gatewayip_array=gatewayip.split('.')
        print(f"gatewayip_array: {gatewayip_array}")
        comport=list(serial.tools.list_ports.comports())[0][0]
        arduino = serial.Serial(port=comport,  baudrate=9600, timeout=.1)
        print(f"arduino: {arduino}")
        arduino_write=dict(
            gatewayip=gatewayip,
            gatewayip_3=gatewayip_array[3],
            gatewayip_2=gatewayip_array[2],
            gatewayip_1=gatewayip_array[1],
            gatewayip_0=gatewayip_array[0],
        )
        arduino_write_dumps=json.dumps(arduino_write)
        print(f"arduino_write_dumps:{arduino_write_dumps}, type:{type(arduino_write_dumps)}")
        # arduino_write_str=str(arduino_write)
        # print(f"arduino_write:{arduino_write_str}")
        # arduino.write(bytes("1",  'utf-8')) #butun config göster
        arduino.write(bytes(arduino_write_dumps,  'utf-8')) #butun config göster
        # arduino.write(bytes("id",  'utf-8')) #butun config göster
        time.sleep(3)    ######## BU SATIR OLMAZSA KESINLIKLE HATA VERIYOR!!!!!!!!!!!!!!!!!!!!!!!
        data = arduino.readline().decode('utf_8')
        print(f"arduino.readline():{data}")
        return  data
        # return arduino_write_str
        # return arduino_write_dumps
        # return  "sabit return data"
        # return  '{"CihazId":1,"CihazAdi":"M(t","CihazPort":90,"CihazIp":"(IP unset)","ServerIp":"192.168.43.130","AgGecidi":"192.168.1.1"}'
    except :
        print("FileNotFoundError exception oluştu")
        return "Exception Hata..."

def write_read_ssid(ssid):
    try:
        ssid=ssid
        # ssid="aaaaaaaa"
        print(f"ssid: {ssid}")
        comport=list(serial.tools.list_ports.comports())[0][0]
        arduino = serial.Serial(port=comport,  baudrate=9600, timeout=.1)
        print(f"arduino: {arduino}")
        arduino_write=dict(
            SSID=ssid,
        )
        arduino_write_dumps=json.dumps(arduino_write)
        print(f"arduino_write_dumps:{arduino_write_dumps}, type:{type(arduino_write_dumps)}")
        arduino.write(bytes(arduino_write_dumps,  'utf-8')) #butun config göster
        time.sleep(3)    ######## BU SATIR OLMAZSA KESINLIKLE HATA VERIYOR!!!!!!!!!!!!!!!!!!!!!!!
        data = arduino.readline().decode('utf_8')
        print(f"arduino.readline():{data}")
        return  data
    except :
        print("FileNotFoundError exception oluştu")
        return "Exception Hata..."

def write_read_password(password):
    try:
        password=password
        # password="aaaaaaaa"
        print(f"password: {password}")
        comport=list(serial.tools.list_ports.comports())[0][0]
        arduino = serial.Serial(port=comport,  baudrate=9600, timeout=.1)
        print(f"arduino: {arduino}")
        arduino_write=dict(
            password=password,
        )
        arduino_write_dumps=json.dumps(arduino_write)
        print(f"arduino_write_dumps:{arduino_write_dumps}, type:{type(arduino_write_dumps)}")
        arduino.write(bytes(arduino_write_dumps,  'utf-8')) #butun config göster
        time.sleep(3)    ######## BU SATIR OLMAZSA KESINLIKLE HATA VERIYOR!!!!!!!!!!!!!!!!!!!!!!!
        data = arduino.readline().decode('utf_8')
        print(f"arduino.readline():{data}")
        return  data
    except :
        print("FileNotFoundError exception oluştu")
        return "Exception Hata..."

def write_read_cihazport(cihazport):
    try:
        cihazport=cihazport
        # cihazport="aaaaaaaa"
        print(f"cihazport: {cihazport}")
        comport=list(serial.tools.list_ports.comports())[0][0]
        arduino = serial.Serial(port=comport,  baudrate=9600, timeout=.1)
        print(f"arduino: {arduino}")
        arduino_write=dict(
            cihazport=cihazport,
        )
        arduino_write_dumps=json.dumps(arduino_write)
        print(f"arduino_write_dumps:{arduino_write_dumps}, type:{type(arduino_write_dumps)}")
        arduino.write(bytes(arduino_write_dumps,  'utf-8')) #butun config göster
        time.sleep(3)    ######## BU SATIR OLMAZSA KESINLIKLE HATA VERIYOR!!!!!!!!!!!!!!!!!!!!!!!
        data = arduino.readline().decode('utf_8')
        print(f"arduino.readline():{data}")
        arduino.close()
        return  data
    except :
        print("FileNotFoundError exception oluştu")
        return "Exception Hata..."
def write_read_reset():
    comport=list(serial.tools.list_ports.comports())[0][0]
    arduino = serial.Serial(port=comport,  baudrate=9600, timeout=.1)
    arduino_write=dict(
        reset="reset"
    )
    arduino_write_dumps=json.dumps(arduino_write)
    print(f"arduino_write_dumps:{arduino_write_dumps}, type:{type(arduino_write_dumps)}")
    arduino.write(bytes(arduino_write_dumps,  'utf-8')) #butun config göster
    time.sleep(3)    ######## BU SATIR OLMAZSA KESINLIKLE HATA VERIYOR!!!!!!!!!!!!!!!!!!!!!!!
    data = arduino.readline().decode('utf_8')
    print(f"arduino.readline():{data}")
    return  data  

# def additional_text(request,id,additional_text):
def additional_text(request):
    additional_text=request.GET.get('additional_text')
    print(f"additional text 1: {additional_text}")
    temp_id=int(request.GET.get('temp_id'))
    temp_user=request.GET.get('temp_user')
    template_name=request.GET.get('temp_template')
    device_name=request.GET.get('device_name')
    device_port=request.GET.get('device_port')
    tarih=timezone.now()
    tarih2=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    tarih3=datetime.timestamp(tarih)
    print(f"tarih1:{tarih},tarih2:{tarih2}")
    edited_item=Temperature.objects.get(id=temp_id)
    edited_item.additionalText=f"{additional_text},ID:{temp_id}, Kullanıcı:{temp_user}, Tarih:{tarih2}"
    print(f"edited_item.additionalText: {edited_item.additionalText}")
    edited_item.save()

    if template_name=="device.html":
        return redirect(f"/app_monitor/cihazlar/{device_name}/port={device_port}")
    else:
        return redirect('/app_monitor')

def additional_text_sil(request,id):
    print(f"additonal_text_sil girdi.")
    edited_item=Temperature.objects.get(id=int(id))
    edited_item.additionalText=""
    edited_item.save()
    print(f"edited_item:{edited_item}")

    return redirect('/app_monitor')

def additional_text_singledevice(request):
    additional_text=request.GET.get('additional_text')
    print(f"additional text 1: {additional_text}")
    temp_id=int(request.GET.get('temp_id'))
    temp_user=request.GET.get('temp_user')
    device_name=request.GET.get('device_name')
    device_port=request.GET.get('device_port')
    tarih=timezone.now()
    tarih2=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    tarih3=datetime.timestamp(tarih)
    print(f"tarih1:{tarih},tarih2:{tarih2}")
    edited_item=Temperature.objects.get(id=temp_id)
    edited_item.additionalText=f"{additional_text},ID:{temp_id}, Kullanıcı:{temp_user}, Tarih:{tarih2}"
    print(f"edited_item.additionalText: {edited_item.additionalText}")
    edited_item.save()

    return redirect(f"/app_monitor/{device_name}/{device_port}")

def additional_text_sil_singledevice(request,id):
    print(f"additonal_text_sil girdi.")
    edited_item=Temperature.objects.get(id=int(id))
    edited_item.additionalText=""
    edited_item.save()
    print(f"edited_item:{edited_item}")

    return redirect('/app_monitor')

def cron_task(request):

    print(f"cron task çalıştı: {timezone.now()}")

def scheduler_cihaz(request):
    datetime_now=datetime.now()
    devices_all=Device.objects.all()
    events=Event.objects.all()[::-1]

    for device in devices_all:
        pass
    
    events=Event.objects.all()[::-1]
    events_total_count=Event.objects.all().count()
    events_alarm1_count=Event.objects.filter(alarm_id=1).filter(event_active=True).count()
    events_alarm2_count=Event.objects.filter(alarm_id=2).filter(event_active=True).count()
    events_clear_count=Event.objects.filter(event_active=False).count()

    context=dict(
        events=events,
        events_total_count=events_total_count,
        events_alarm1_count=events_alarm1_count,
        events_alarm2_count=events_alarm2_count,
        events_clear_count=events_clear_count,
        datetime_now=datetime_now,
    )
    return render(request,"app_monitor/events.html",context)

def event_list_view(request):
    events=Event.objects.all()[::-1]

    events_total_state=request.GET.get("events_total_state")
    events_alarm1_state=request.GET.get("events_alarm1_state")
    events_alarm2_state=request.GET.get("events_alarm2_state")
    events_alarm_clear_state=request.GET.get("events_alarm_clear_state")

    print(f"events_alarm1_state:{events_alarm1_state},events_alarm2_state:{events_alarm2_state}, events_alarm_clear_state:{events_alarm_clear_state}")

    if events_alarm1_state == "true":
        # events_all=Event.objects.filter(alarm_id=1).filter(event_active=True)[::-1]
        # events_1=Event.objects.filter(alarm_id=1)
        events_all=[]
    else:
        # events_all=Event.objects.exclude(alarm_id=1).filter(event_active=True)[::-1]
        events_all=Event.objects.filter(alarm_id=1).filter(event_active=True)[::-1]
        # events_all=[]
    if events_alarm2_state == "true":
        # events_all=Event.objects.filter(alarm_id=2).filter(event_active=True)[::-1]
        # events_2=Event.objects.filter(alarm_id=2).filter(event_active=True)
        # events_all=sorted(chain(events_2,events_all),key=attrgetter('id'),reverse=True)
        # events_all=list(chain(events_2,events_all),key=attrgetter('id'),reverse=True)
        # events_2=list(chain(events_2,events_1))
        # events_2=events_1+events_2
        pass
    else:
        events_2=Event.objects.filter(alarm_id=2).filter(event_active=True)
        events_all=sorted(chain(events_2,events_all),key=attrgetter('id'),reverse=True)
        # events_all=events_all.exclude(alarm_id=2).filter(event_active=True)[::-1]
        # events_all=Event.objects.all()[::-1]
        
    if events_alarm_clear_state == "true":
    # if events_alarm_clear_state == True: # request ten gelen değerle kontrol yapıldığı için text olarak true değeri gelmekte.
        # events_all=Event.objects.filter(event_active=False)[::-1]
        # events_all=Event.objects.filter(event_active=False)[::-1]
        # events_clear=Event.objects.filter(event_active=False)
        # events_all=sorted(chain(events_clear,events_all),key=attrgetter('id'),reverse=True)
        # events_all=events_clear+events_2
        pass
    else:
        events_clear=Event.objects.filter(event_active=False)
        events_all=sorted(chain(events_clear,events_all),key=attrgetter('id'),reverse=True)
        # events_all=Event.objects.all()[::-1]
        # events_all=events_all.exclude(event_active=True)[::-1]

    # if (events_alarm1_state=="false") and (events_alarm2_state=="false") and (events_alarm_clear_state=="false"):
    # # if (events_alarm1_state==False) and (events_alarm2_state==False) and (events_alarm_clear_state==False):
    #     events_all=Event.objects.all()[::-1]

    # if events_total_state == "true":
    #     events_all=Event.objects.all()[::-1]
    # else:
    #     events_all=[]

    # datetime_now=datetime.now()
    datetime_now=timezone.now()
    context=dict(
        events=events,
        events_all=events_all,
        datetime_now=datetime_now,
    )
    print(f"event_list_view girdi, time:{datetime_now}")
    return render(request,"app_monitor/events_ajax.html",context)

#Excel export ALL
def export_to_excel_event_all(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="events_all.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "ALL EVENTS"

    # Add headers
    headers = ["id","DEVICE NAME","DEVICE ID","ALARM ID","ALARM NAME", "EVENT ACTIVE", "START TIME","FINISH TIME","INFO"]
    ws.append(headers)

    events = Event.objects.all()[::-1] 
        # temp = Temperature.objects.filter(device_name=cihazadi)
    for event in events:
        try:
            ws.append([event.id,event.device_name,event.device_id.device_id,event.alarm_id,event.alarm_name, event.event_active, event.start_time,event.finish_time,event.info])
            # ws.append([temps.id,temps.device_name,temps.temperature, temps.humidity, temps.volcum])
        except AttributeError:
            print("AttributeError oluştu...")

    # Save the workbook to the HttpResponse
    wb.save(response)
    return response

def event_buton_view(request):
    all_events=request.GET.get("all_events")
    critical_events=request.GET.get("critical_events")
    info_events=request.GET.get("info_events")

    # all_events_query=